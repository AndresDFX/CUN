# -*- coding: utf-8 -*-
r"""Lee el CRONOGRAMA DE SUSTENTACIÓN del periodo y saca los proyectos que me toca evaluar.

El cronograma institucional es un `.xlsx` con una hoja `CRONOGRAMA` (quién evalúa
qué especialización, en qué rol y qué día) y una hoja por especialización con el
detalle de cada grupo: integrantes, proyecto, línea, objetivo, horario y un
**hipervínculo de Drive** a la carpeta con el trabajo de grado.

Dos cosas que no son evidentes y cuestan un rato:

1. **Las tablas empiezan en la columna B, no en la A.** La columna A está vacía y
   sirve de margen. Leer `row[0]` devuelve cadenas vacías y el parser cree que la
   hoja no tiene grupos.
2. **La carpeta de cada grupo no hay que descargarla.** El hipervínculo apunta a
   un id de Drive que Drive para escritorio ya monta en
   `G:\.shortcut-targets-by-id\<id>\`. `Get-ChildItem` sobre el padre no siempre
   los lista (son accesos directos bajo demanda), pero la ruta directa existe y
   se puede leer con cualquier herramienta local.

Uso:
    python config/evaluaciones/proyectos_sustentacion.py
    python config/evaluaciones/proyectos_sustentacion.py --correo julian_castanoe@cun.edu.co
    python config/evaluaciones/proyectos_sustentacion.py --json salida.json
    python config/evaluaciones/proyectos_sustentacion.py --xlsx "otra/ruta.xlsx" --hoja MARIA-ESPTD
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys

sys.stdout.reconfigure(encoding="utf-8")

RAIZ = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
EVALUACIONES = os.path.join(RAIZ, "Especializacion", "Evaluaciones")
MONTURA = r"G:\.shortcut-targets-by-id"
CORREO_POR_DEFECTO = "julian_castanoe@cun.edu.co"

# La tabla de detalle arranca en la columna B: desplazamiento de 1 sobre el índice 0.
COL = {"grupo": 1, "no": 2, "estudiante": 3, "cedula": 4, "correo": 5,
       "proyecto": 6, "linea": 7, "objetivo": 8, "observaciones": 9, "horario": 10}


def _t(celda) -> str:
    return "" if celda.value is None else re.sub(r"\s+", " ", str(celda.value)).strip()


def cronograma_mas_reciente() -> str:
    """El .xlsx de sustentación del periodo más nuevo que haya en Evaluaciones/."""
    cands = []
    for raiz, _, ficheros in os.walk(EVALUACIONES):
        for f in ficheros:
            if f.lower().endswith(".xlsx") and "cronograma" in f.lower() and not f.startswith("~$"):
                cands.append(os.path.join(raiz, f))
    if not cands:
        raise SystemExit(f"no encontré ningún cronograma .xlsx bajo {EVALUACIONES}")
    return max(cands, key=lambda p: os.path.getmtime(p))


def mis_asignaciones(wb, correo: str) -> list[dict]:
    """Filas de la hoja CRONOGRAMA donde aparece mi correo, con el rol que tengo."""
    ws = wb["CRONOGRAMA"]
    roles = {}   # columna del correo -> rol
    encabezado = None
    for fila in ws.iter_rows():
        vals = [_t(c) for c in fila]
        if "ESPECIALIZACIÓN" in vals and "GRUPOS" in [v.upper() for v in vals]:
            encabezado = vals
            for i, v in enumerate(vals):
                if v.upper() == "CORREO" and i > 0:
                    roles[i] = vals[i - 1]  # el nombre va justo antes del correo
            break
    if not encabezado:
        raise SystemExit("la hoja CRONOGRAMA no tiene la fila de encabezado esperada")

    fuera, esp_actual = [], ""
    for fila in ws.iter_rows():
        vals = [_t(c) for c in fila]
        if not any(vals):
            continue
        if vals[1].upper().startswith("ESPECIALIZACIÓN EN"):
            esp_actual = vals[1]
        mios = [roles[i] for i in roles if i < len(vals) and vals[i].lower() == correo.lower()]
        if mios:
            hoja = ""
            for v in vals:
                m = re.search(r"'?([A-ZÁÉÍÓÚÑ]+-[A-Z]+)'?!", v)
                if m:
                    hoja = m.group(1)
            fuera.append({"especializacion": esp_actual, "rol": mios[0],
                          "hoja": hoja, "fila": [v for v in vals if v]})
    return fuera


def grupos_de_hoja(wb, hoja: str) -> list[dict]:
    ws = wb[hoja]
    enlaces = {c.coordinate: (c.hyperlink.target or c.hyperlink.location)
               for f in ws.iter_rows() for c in f if c.hyperlink is not None}
    grupos, actual, dia = [], None, ""
    for fila in ws.iter_rows():
        a = _t(fila[COL["grupo"]]) if len(fila) > COL["grupo"] else ""
        alto = a.upper()
        if alto.startswith("SUSTENTACIÓN DÍA") or alto.startswith("SUSTENTACION DÍA"):
            dia = a
            continue
        if alto.startswith("CÓDIGO GRUPO") or alto.startswith("CODIGO GRUPO"):
            continue
        if re.match(r"^\d{2}[A-Z]{2}\d-G-\d+", a):
            actual = {k: (_t(fila[i]) if len(fila) > i else "") for k, i in COL.items()}
            actual.update({"dia": dia, "integrantes": [],
                           "hipervinculo": enlaces.get(fila[COL["grupo"]].coordinate, "")})
            actual["drive_id"] = _id_drive(actual["hipervinculo"])
            actual["carpeta"] = _carpeta(actual["drive_id"])
            actual["documentos"] = _documentos(actual["carpeta"])
            grupos.append(actual)
        if actual is not None and len(fila) > COL["correo"] and _t(fila[COL["estudiante"]]):
            actual["integrantes"].append({
                "nombre": _t(fila[COL["estudiante"]]),
                "cedula": _t(fila[COL["cedula"]]).replace(".0", ""),
                "correo": _t(fila[COL["correo"]]),
            })
    return grupos


def _id_drive(url: str) -> str:
    m = re.search(r"(?:[?&]id=|/d/|/folders/)([A-Za-z0-9_-]{20,})", url or "")
    return m.group(1) if m else ""


def _carpeta(drive_id: str) -> str:
    if not drive_id:
        return ""
    p = os.path.join(MONTURA, drive_id)
    if not os.path.isdir(p):
        return ""
    # Drive suele montar el id con una única subcarpeta dentro, que es la del grupo.
    hijos = [h for h in os.listdir(p) if os.path.isdir(os.path.join(p, h))]
    return os.path.join(p, hijos[0]) if len(hijos) == 1 else p


def _documentos(carpeta: str) -> list[dict]:
    if not carpeta or not os.path.isdir(carpeta):
        return []
    fuera = []
    for raiz, _, ficheros in os.walk(carpeta):
        for f in sorted(ficheros):
            if f == "desktop.ini" or f.startswith("~$"):
                continue
            ruta = os.path.join(raiz, f)
            fuera.append({"nombre": f, "ruta": ruta, "bytes": os.path.getsize(ruta)})
    return fuera


def main() -> None:
    ap = argparse.ArgumentParser(description="Proyectos de grado que me toca evaluar este periodo.")
    ap.add_argument("--correo", default=CORREO_POR_DEFECTO)
    ap.add_argument("--xlsx", default="")
    ap.add_argument("--hoja", default="", help="forzar una hoja de detalle concreta")
    ap.add_argument("--json", default="", help="volcar el resultado a este archivo")
    a = ap.parse_args()

    from openpyxl import load_workbook

    ruta = a.xlsx or cronograma_mas_reciente()
    print(f"### Cronograma: {ruta}")
    wb = load_workbook(ruta, data_only=True)

    asigs = mis_asignaciones(wb, a.correo)
    if not asigs:
        print(f"!! {a.correo} no aparece en la hoja CRONOGRAMA. Hojas: {wb.sheetnames}")
    hojas = [a.hoja] if a.hoja else [x["hoja"] for x in asigs if x["hoja"] in wb.sheetnames]

    for x in asigs:
        print(f"\n## {x['especializacion']}\n   MI ROL: {x['rol']}   hoja de detalle: {x['hoja'] or '(no declarada)'}")
        print(f"   fila: {' | '.join(x['fila'])[:300]}")

    salida = {"cronograma": ruta, "correo": a.correo, "asignaciones": asigs, "grupos": []}
    for hoja in dict.fromkeys(hojas):
        grupos = grupos_de_hoja(wb, hoja)
        salida["grupos"].extend(grupos)
        print(f"\n### Hoja {hoja}: {len(grupos)} grupos")
        for g in grupos:
            faltan = "  ⚠ SIN DOCUMENTOS" if not g["documentos"] else ""
            print(f"\n[{g['grupo']}] {g['dia']}  {g['horario'] or '(sin horario)'}{faltan}")
            print(f"   {g['proyecto'][:150]}")
            print(f"   línea: {g['linea']}   integrantes: {len(g['integrantes'])}")
            if not g["carpeta"]:
                print(f"   carpeta NO montada (id {g['drive_id'] or 'sin enlace'})")
            for d in g["documentos"]:
                print(f"   - {d['nombre']}  ({d['bytes']:,} B)")

    if a.json:
        with open(a.json, "w", encoding="utf-8") as f:
            json.dump(salida, f, ensure_ascii=False, indent=2)
        print(f"\n-> {a.json}")


if __name__ == "__main__":
    main()
